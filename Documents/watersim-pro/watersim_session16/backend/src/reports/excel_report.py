"""
WaterSim Pro — Excel Report Generator
Reads JSON (single run or comparison array) from stdin, writes .xlsx to stdout.

Usage:
  echo '{"mode":"single","data":{...}}' | python3 excel_report.py
  echo '{"mode":"comparison","runs":[...]}' | python3 excel_report.py
"""
import sys
import json
import io
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.stderr.write("openpyxl not installed — run: pip install openpyxl\n")
    sys.exit(1)

# ── Palette ────────────────────────────────────────────────────────────────────
NAVY      = "1F3864"
BRAND     = "2E75B6"
BRAND_LT  = "DEEAF1"
EMERALD   = "1A7A4A"
EME_LT    = "D6F0E2"
RED       = "C00000"
RED_LT    = "FFDEDE"
AMBER     = "C55A11"
AMB_LT    = "FCE4D6"
GRAY_HDR  = "F2F2F2"
GRAY_ALT  = "F8F9FA"
WHITE     = "FFFFFF"
BLACK     = "000000"

def mk_font(bold=False, color=BLACK, size=10, name="Arial"):
    return Font(name=name, size=size, bold=bold, color=color)

def mk_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def mk_border(style="thin"):
    s = Side(style=style, color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def mk_align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def fmt_num(v, dec=2):
    if v is None: return "—"
    try:
        n = float(v)
        return round(n, dec)
    except (TypeError, ValueError):
        return "—"

def fmt_pct(inf, eff):
    try:
        i, e = float(inf), float(eff)
        if i == 0: return "—"
        return round((i - e) / i * 100, 1)
    except (TypeError, ValueError):
        return "—"

def style_header_row(ws, row, col_start, col_end, fill_hex, font_color=WHITE):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = mk_font(bold=True, color=font_color, size=10)
        cell.fill = mk_fill(fill_hex)
        cell.alignment = mk_align("center")
        cell.border = mk_border()

def style_data_row(ws, row, col_start, col_end, alt=False):
    fill = mk_fill(GRAY_ALT if alt else WHITE)
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.border = mk_border()
        cell.alignment = mk_align()

def set_col_widths(ws, widths):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

# ── Quality params config ──────────────────────────────────────────────────────
QUALITY_PARAMS = [
    {"key": "Q",    "label": "Flow",                    "unit": "m³/d",  "dec": 0},
    {"key": "BOD",  "label": "BOD₅",                   "unit": "mg/L",  "dec": 1},
    {"key": "COD",  "label": "COD",                     "unit": "mg/L",  "dec": 1},
    {"key": "TSS",  "label": "Total Suspended Solids",  "unit": "mg/L",  "dec": 1},
    {"key": "TN",   "label": "Total Nitrogen",          "unit": "mg/L",  "dec": 2},
    {"key": "NH4",  "label": "Ammonia (NH₄-N)",         "unit": "mg/L",  "dec": 2},
    {"key": "NO3",  "label": "Nitrate (NO₃-N)",         "unit": "mg/L",  "dec": 2},
    {"key": "TP",   "label": "Total Phosphorus",        "unit": "mg/L",  "dec": 2},
    {"key": "DO",   "label": "Dissolved Oxygen",        "unit": "mg/L",  "dec": 2},
    {"key": "pH",   "label": "pH",                      "unit": "S.U.", "dec": 2},
    {"key": "temp", "label": "Temperature",             "unit": "°C",    "dec": 1},
]

# ════════════════════════════════════════════════════════════════════════════════
# PLAIN-LANGUAGE SHEET
# ════════════════════════════════════════════════════════════════════════════════

def build_plain_sheet(wb, data):
    """Prepend a 'Plain summary' sheet built from the backend `plain` key.

    Skips cleanly when `plain` is absent (old cached report JSON) or malformed.
    """
    plain = data.get("plain")
    if not isinstance(plain, dict):
        return
    try:
        ws = wb.create_sheet("Plain summary", 0)
        ws.sheet_view.showGridLines = False

        # Title band
        ws.merge_cells("A1:E1")
        c = ws["A1"]
        c.value = "In plain words — what this simulation means"
        c.font = mk_font(bold=True, color=WHITE, size=13)
        c.fill = mk_fill(NAVY)
        c.alignment = mk_align("center")
        ws.row_dimensions[1].height = 26

        # Verdict banner
        verdict = plain.get("verdict") or {}
        status = verdict.get("status")
        r = 3
        ws.merge_cells(f"A{r}:E{r}")
        vc = ws.cell(r, 1)
        vc.value = verdict.get("headline") or "No plain-language verdict available."
        if status == "pass":
            vc.fill = mk_fill(EME_LT); vc.font = mk_font(bold=True, color=EMERALD, size=11)
        elif status == "fail":
            vc.fill = mk_fill(RED_LT); vc.font = mk_font(bold=True, color=RED, size=11)
        else:
            vc.fill = mk_fill(GRAY_HDR); vc.font = mk_font(bold=True, size=11)
        vc.alignment = mk_align("left", wrap=True)
        ws.row_dimensions[r].height = 30
        r += 1
        if verdict.get("detail"):
            ws.merge_cells(f"A{r}:E{r}")
            dc = ws.cell(r, 1)
            dc.value = verdict["detail"]
            dc.font = mk_font(size=9)
            dc.alignment = mk_align("left", wrap=True)
            ws.row_dimensions[r].height = 26
            r += 1
        r += 1

        # Water story
        for item in (plain.get("waterStory") or []):
            if not isinstance(item, dict):
                continue
            ws.merge_cells(f"A{r}:E{r}")
            wc = ws.cell(r, 1)
            wc.value = f"{item.get('label', '')}: {item.get('text', '')}"
            wc.font = mk_font(size=9)
            wc.alignment = mk_align("left", wrap=True)
            ws.row_dimensions[r].height = 26
            r += 1
        r += 1

        # Quality rows
        q_rows = [q for q in (plain.get("qualityRows") or []) if isinstance(q, dict)]
        if q_rows:
            ws.merge_cells(f"A{r}:E{r}")
            ws.cell(r, 1).value = "HOW CLEAN IS THE WATER?"
            ws.cell(r, 1).font = mk_font(bold=True, color=WHITE, size=10)
            ws.cell(r, 1).fill = mk_fill(BRAND)
            ws.cell(r, 1).alignment = mk_align("center")
            r += 1
            headers = ["What we measure", "Coming in (mg/L)", "Going out (mg/L)", "Removed (%)", "Verdict"]
            for ci, h in enumerate(headers, 1):
                ws.cell(r, ci).value = h
            style_header_row(ws, r, 1, 5, BRAND)
            r += 1
            for qi, q in enumerate(q_rows):
                style_data_row(ws, r, 1, 5, qi % 2 == 1)
                judgment = q.get("judgment")
                verdict_txt = {"good": "✓ good", "ok": "~ OK", "poor": "✗ poor"}.get(judgment, "—")
                vals = [q.get("friendly", q.get("param", "")),
                        fmt_num(q.get("in"), 1), fmt_num(q.get("out"), 1),
                        fmt_num(q.get("removalPct"), 1), verdict_txt]
                for ci, v in enumerate(vals, 1):
                    cell = ws.cell(r, ci)
                    cell.value = v
                    if ci in (2, 3, 4):
                        cell.alignment = mk_align("right")
                if judgment == "good":
                    ws.cell(r, 5).font = mk_font(bold=True, color=EMERALD)
                elif judgment == "poor":
                    ws.cell(r, 5).font = mk_font(bold=True, color=RED)
                elif judgment == "ok":
                    ws.cell(r, 5).font = mk_font(bold=True, color=AMBER)
                r += 1
            r += 1

        # Compliance lines
        c_story = [c2 for c2 in (plain.get("complianceStory") or []) if isinstance(c2, dict)]
        if c_story:
            ws.merge_cells(f"A{r}:E{r}")
            ws.cell(r, 1).value = "IS THE WATER LEGAL TO RELEASE?"
            ws.cell(r, 1).font = mk_font(bold=True, color=WHITE, size=10)
            ws.cell(r, 1).fill = mk_fill(BRAND)
            ws.cell(r, 1).alignment = mk_align("center")
            r += 1
            for c2 in c_story:
                ws.merge_cells(f"A{r}:E{r}")
                lc = ws.cell(r, 1)
                lc.value = f"• {c2.get('text', '')}"
                sev = c2.get("severity")
                if sev == "none":
                    lc.font = mk_font(color=EMERALD, size=9)
                elif sev in ("high", "medium"):
                    lc.font = mk_font(color=RED, size=9)
                else:
                    lc.font = mk_font(color=AMBER, size=9)
                lc.alignment = mk_align("left", wrap=True)
                ws.row_dimensions[r].height = 26
                r += 1
            r += 1

        # Treatment steps
        steps = [s for s in (plain.get("treatmentSteps") or []) if isinstance(s, dict)]
        if steps:
            ws.merge_cells(f"A{r}:E{r}")
            ws.cell(r, 1).value = "THE JOURNEY, STEP BY STEP"
            ws.cell(r, 1).font = mk_font(bold=True, color=WHITE, size=10)
            ws.cell(r, 1).fill = mk_fill(BRAND)
            ws.cell(r, 1).alignment = mk_align("center")
            r += 1
            for i, s in enumerate(steps, 1):
                ws.merge_cells(f"A{r}:E{r}")
                sc = ws.cell(r, 1)
                txt = f"{i}. {s.get('label', 'Step')} — {s.get('explanation', '')}"
                if s.get("keyFact"):
                    txt += f"  ({s['keyFact']})"
                sc.value = txt
                sc.font = mk_font(size=9)
                sc.alignment = mk_align("left", wrap=True)
                ws.row_dimensions[r].height = 26
                r += 1
            r += 1

        # Cost lines
        cost_lines = (plain.get("costStory") or {}).get("lines") or []
        if cost_lines:
            ws.merge_cells(f"A{r}:E{r}")
            ws.cell(r, 1).value = "WHAT IT COSTS"
            ws.cell(r, 1).font = mk_font(bold=True, color=WHITE, size=10)
            ws.cell(r, 1).fill = mk_fill(BRAND)
            ws.cell(r, 1).alignment = mk_align("center")
            r += 1
            for line in cost_lines:
                ws.merge_cells(f"A{r}:E{r}")
                lc = ws.cell(r, 1)
                lc.value = f"• {line}"
                lc.font = mk_font(size=9)
                lc.alignment = mk_align("left", wrap=True)
                ws.row_dimensions[r].height = 22
                r += 1
            r += 1

        # Glossary
        glossary = [g for g in (plain.get("glossary") or []) if isinstance(g, dict)]
        if glossary:
            ws.merge_cells(f"A{r}:E{r}")
            ws.cell(r, 1).value = "PLAIN-WORDS DICTIONARY"
            ws.cell(r, 1).font = mk_font(bold=True, color=WHITE, size=10)
            ws.cell(r, 1).fill = mk_fill(EMERALD)
            ws.cell(r, 1).alignment = mk_align("center")
            r += 1
            for gi, g in enumerate(glossary):
                style_data_row(ws, r, 1, 5, gi % 2 == 1)
                ws.cell(r, 1).value = g.get("term", "")
                ws.cell(r, 1).font = mk_font(bold=True, size=9)
                ws.merge_cells(f"B{r}:E{r}")
                gc = ws.cell(r, 2)
                gc.value = g.get("definition", "")
                gc.font = mk_font(size=9)
                gc.alignment = mk_align("left", wrap=True)
                r += 1

        set_col_widths(ws, {"A": 34, "B": 18, "C": 18, "D": 14, "E": 12})
    except Exception:
        # Never let the plain layer break the engineering workbook: drop the
        # partial sheet if anything went wrong.
        try:
            if "Plain summary" in wb.sheetnames:
                del wb["Plain summary"]
        except Exception:
            pass


# ════════════════════════════════════════════════════════════════════════════════
# SINGLE RUN REPORT
# ════════════════════════════════════════════════════════════════════════════════

def build_single(wb, data):
    results = data.get("results", {})
    summary = results.get("summary", {})
    inf = summary.get("influent", {})
    eff = summary.get("effluent", {})
    limits = results.get("permitLimitsUsed") or {}
    cost = results.get("costBreakdown")
    unit_results = results.get("unitResults", {})
    warnings = data.get("warnings", [])

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    # Title band
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "WaterSim Pro — Simulation Report"
    c.font = mk_font(bold=True, color=WHITE, size=14)
    c.fill = mk_fill(NAVY)
    c.alignment = mk_align("center")
    ws.row_dimensions[1].height = 28

    # Metadata block
    meta = [
        ("Project",      data.get("project_name", "—")),
        ("Flowsheet",    data.get("flowsheet_name", "—")),
        ("Run ID",       data.get("run_id", "—")),
        ("Mode",         (data.get("mode") or "—").replace("_", " ").title()),
        ("Completed",    data.get("completed_at", "—")[:19] if data.get("completed_at") else "—"),
        ("Created by",   data.get("created_by", "—")),
    ]
    for i, (k, v) in enumerate(meta):
        r = 2 + i
        ws.cell(r, 1).value = k
        ws.cell(r, 1).font = mk_font(bold=True, size=9)
        ws.cell(r, 1).fill = mk_fill(GRAY_HDR)
        ws.cell(r, 2).value = v
        ws.cell(r, 2).font = mk_font(size=9)

    # Compliance banner
    compliant = summary.get("compliant")
    violations = summary.get("permit_violations", [])
    r_comp = 9
    ws.merge_cells(f"A{r_comp}:G{r_comp}")
    cc = ws.cell(r_comp, 1)
    if compliant is True:
        cc.value = "✓  ALL PERMIT LIMITS MET"
        cc.fill = mk_fill(EME_LT)
        cc.font = mk_font(bold=True, color=EMERALD, size=11)
    elif compliant is False:
        cc.value = f"✗  {len(violations)} PERMIT VIOLATION(S)"
        cc.fill = mk_fill(RED_LT)
        cc.font = mk_font(bold=True, color=RED, size=11)
    else:
        cc.value = "—  COMPLIANCE UNKNOWN"
        cc.fill = mk_fill(GRAY_HDR)
        cc.font = mk_font(bold=True, size=11)
    cc.alignment = mk_align("center")
    ws.row_dimensions[r_comp].height = 22

    # Violations list
    r = r_comp + 1
    for v in violations:
        ws.merge_cells(f"A{r}:G{r}")
        wc = ws.cell(r, 1)
        wc.value = f"   • {v}"
        wc.font = mk_font(color=RED, size=9)
        wc.fill = mk_fill(RED_LT)
        r += 1

    # ── Effluent Quality Table ─────────────────────────────────────────────────
    r += 1
    ws.merge_cells(f"A{r}:G{r}")
    ws.cell(r, 1).value = "INFLUENT & EFFLUENT QUALITY"
    ws.cell(r, 1).font = mk_font(bold=True, color=WHITE, size=10)
    ws.cell(r, 1).fill = mk_fill(BRAND)
    ws.cell(r, 1).alignment = mk_align("center")
    r += 1

    headers = ["Parameter", "Unit", "Influent", "Effluent", "Removal (%)", "Permit Limit", "Status"]
    for ci, h in enumerate(headers, 1):
        ws.cell(r, ci).value = h
    style_header_row(ws, r, 1, 7, BRAND)
    r += 1

    for qi, p in enumerate(QUALITY_PARAMS):
        inf_v = fmt_num(inf.get(p["key"]), p["dec"])
        eff_v = fmt_num(eff.get(p["key"]), p["dec"])
        rem_v = fmt_pct(inf.get(p["key"]), eff.get(p["key"]))
        lim_v = fmt_num(limits.get(p["key"]), p["dec"]) if limits.get(p["key"]) is not None else "—"

        # Status
        if (limits.get(p["key"]) is not None and
                eff.get(p["key"]) is not None and
                p["key"] not in ("Q", "pH", "temp", "DO")):
            status_pass = float(eff.get(p["key"]) or 0) <= float(limits.get(p["key"]) or 999)
            status = "PASS" if status_pass else "FAIL"
        else:
            status = "—"

        row_data = [p["label"], p["unit"], inf_v, eff_v, rem_v, lim_v, status]
        alt = qi % 2 == 1
        style_data_row(ws, r, 1, 7, alt)
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(r, ci)
            cell.value = val
            if ci in (3, 4, 5, 6):
                cell.alignment = mk_align("right")
            if ci == 7 and status == "PASS":
                cell.font = mk_font(bold=True, color=EMERALD)
            elif ci == 7 and status == "FAIL":
                cell.font = mk_font(bold=True, color=RED)
        r += 1

    # ── KPI strip ────────────────────────────────────────────────────────────
    r += 1
    kpis = [
        ("Influent Flow (m³/d)", fmt_num(inf.get("Q"), 0)),
        ("Effluent Flow (m³/d)", fmt_num(eff.get("Q"), 0)),
        ("BOD Removal (%)",      fmt_pct(inf.get("BOD"), eff.get("BOD"))),
        ("TN Removal (%)",       fmt_pct(inf.get("TN"), eff.get("TN"))),
    ]
    if cost:
        kpis.append(("Unit Cost (USD/m³)", fmt_num(cost.get("cost_per_m3_treated_USD"), 3)))

    for ki, (label, val) in enumerate(kpis):
        c_start = 1 + ki * 1
        ws.cell(r, 1 + ki).value = label
        ws.cell(r, 1 + ki).font = mk_font(bold=True, size=9, color=BRAND)
        ws.cell(r + 1, 1 + ki).value = val
        ws.cell(r + 1, 1 + ki).font = mk_font(bold=True, size=12)

    set_col_widths(ws, {"A": 28, "B": 10, "C": 12, "D": 12, "E": 13, "F": 13, "G": 10})

    # ── Sheet 2: Cost Breakdown ───────────────────────────────────────────────
    if cost:
        wc2 = wb.create_sheet("Cost Breakdown")
        wc2.sheet_view.showGridLines = False

        wc2.merge_cells("A1:D1")
        wc2["A1"].value = "Annual Operating Cost Breakdown"
        wc2["A1"].font = mk_font(bold=True, color=WHITE, size=13)
        wc2["A1"].fill = mk_fill(NAVY)
        wc2["A1"].alignment = mk_align("center")
        wc2.row_dimensions[1].height = 26

        cost_sections = [
            ("Energy",      cost.get("energy", {}) or {}),
            ("Chemicals",   cost.get("chemicals", {}) or {}),
            ("Sludge",      cost.get("sludge", {}) or {}),
            ("Labour",      cost.get("labour", {}) or {}),
            ("Maintenance", cost.get("maintenance", {}) or {}),
        ]

        r = 3
        headers = ["Cost Category", "Annual Cost (USD)", "% of Total"]
        for ci, h in enumerate(headers, 1):
            wc2.cell(r, ci).value = h
        style_header_row(wc2, r, 1, 3, BRAND)
        r += 1

        total = cost.get("total_USD_yr", 0) or 0
        for si, (name, sec) in enumerate(cost_sections):
            val = sec.get("cost_USD_yr") or sec.get("total_USD_yr")
            style_data_row(wc2, r, 1, 3, si % 2 == 1)
            wc2.cell(r, 1).value = name
            wc2.cell(r, 2).value = fmt_num(val, 0) if val is not None else "—"
            wc2.cell(r, 2).alignment = mk_align("right")
            if val is not None and total > 0:
                pct = round(float(val) / total * 100, 1)
                wc2.cell(r, 3).value = f"{pct}%"
            wc2.cell(r, 3).alignment = mk_align("right")
            r += 1

        # Total row
        wc2.cell(r, 1).value = "TOTAL"
        wc2.cell(r, 1).font = mk_font(bold=True)
        wc2.cell(r, 2).value = fmt_num(total, 0)
        wc2.cell(r, 2).font = mk_font(bold=True)
        wc2.cell(r, 2).alignment = mk_align("right")
        wc2.cell(r, 3).value = "100.0%"
        wc2.cell(r, 3).alignment = mk_align("right")
        for ci in range(1, 4):
            wc2.cell(r, ci).fill = mk_fill(BRAND_LT)
            wc2.cell(r, ci).border = mk_border()

        r += 2
        wc2.cell(r, 1).value = "Unit Cost (USD/m³)"
        wc2.cell(r, 2).value = fmt_num(cost.get("cost_per_m3_treated_USD"), 3)
        wc2.cell(r + 1, 1).value = "Energy (kWh/yr)"
        wc2.cell(r + 1, 2).value = fmt_num((cost.get("energy") or {}).get("total_kWh_yr"), 0)
        wc2.cell(r + 2, 1).value = "Dry Sludge (t/yr)"
        wc2.cell(r + 2, 2).value = fmt_num((cost.get("sludge") or {}).get("dry_tonnes_yr"), 0)
        wc2.cell(r + 3, 1).value = "Labour (FTE)"
        wc2.cell(r + 3, 2).value = (cost.get("labour") or {}).get("staff_count", "—")

        set_col_widths(wc2, {"A": 24, "B": 20, "C": 14})

    # ── Sheet 3: Unit Operations ──────────────────────────────────────────────
    if unit_results:
        wu = wb.create_sheet("Unit Operations")
        wu.sheet_view.showGridLines = False
        wu.merge_cells("A1:D1")
        wu["A1"].value = "Unit Operation Performance Metrics"
        wu["A1"].font = mk_font(bold=True, color=WHITE, size=13)
        wu["A1"].fill = mk_fill(NAVY)
        wu["A1"].alignment = mk_align("center")

        r = 3
        for node_id, ur in unit_results.items():
            metrics = ur.get("metrics", {})
            if not metrics:
                continue
            op_type = (ur.get("paletteType") or ur.get("type") or "Unknown").replace("_", " ").title()
            wu.merge_cells(f"A{r}:D{r}")
            wu.cell(r, 1).value = f"{op_type} — {node_id}"
            wu.cell(r, 1).font = mk_font(bold=True, color=WHITE, size=9)
            wu.cell(r, 1).fill = mk_fill(BRAND)
            wu.cell(r, 1).alignment = mk_align("center")
            r += 1

            for mi, (k, v) in enumerate(metrics.items()):
                alt = mi % 2 == 1
                style_data_row(wu, r, 1, 4, alt)
                wu.cell(r, 1).value = k.replace("_", " ").title()
                wu.cell(r, 2).value = fmt_num(v) if isinstance(v, (int, float)) else str(v)
                wu.cell(r, 2).alignment = mk_align("right")
                r += 1
            r += 1

        set_col_widths(wu, {"A": 30, "B": 16, "C": 12, "D": 12})

    # ── Sheet 4: Warnings ────────────────────────────────────────────────────
    if warnings:
        ww = wb.create_sheet("Warnings")
        ww["A1"].value = "Simulation Warnings"
        ww["A1"].font = mk_font(bold=True, color=WHITE, size=12)
        ww["A1"].fill = mk_fill(AMBER)
        for i, w in enumerate(warnings):
            ww.cell(2 + i, 1).value = w
            ww.cell(2 + i, 1).font = mk_font(color=AMBER, size=9)
        set_col_widths(ww, {"A": 80})

    # ── Sheet 0: Plain-language summary (inserted first) ─────────────────────
    build_plain_sheet(wb, data)


# ════════════════════════════════════════════════════════════════════════════════
# COMPARISON REPORT
# ════════════════════════════════════════════════════════════════════════════════

def build_comparison(wb, runs):
    n = len(runs)

    # ── Sheet 1: Effluent Quality Comparison ─────────────────────────────────
    ws = wb.active
    ws.title = "Effluent Comparison"
    ws.sheet_view.showGridLines = False

    ws.merge_cells(f"A1:{get_column_letter(2 + n)}1")
    ws["A1"].value = "WaterSim Pro — Scenario Comparison Report"
    ws["A1"].font = mk_font(bold=True, color=WHITE, size=14)
    ws["A1"].fill = mk_fill(NAVY)
    ws["A1"].alignment = mk_align("center")
    ws.row_dimensions[1].height = 28

    # Run headers
    r = 2
    ws.cell(r, 1).value = "Scenario →"
    ws.cell(r, 1).font = mk_font(bold=True, size=9)
    ws.cell(r, 1).fill = mk_fill(GRAY_HDR)
    for i, run in enumerate(runs):
        label = run.get("label") or run.get("flowsheet_name", f"Run {i+1}")
        ws.cell(r, 2 + i).value = label
        ws.cell(r, 2 + i).font = mk_font(bold=True, color=WHITE, size=9)
        ws.cell(r, 2 + i).fill = mk_fill(BRAND)
        ws.cell(r, 2 + i).alignment = mk_align("center")
        ws.cell(r, 2 + i).border = mk_border()
    r += 1

    # Sub-row: project
    ws.cell(r, 1).value = "Project"
    ws.cell(r, 1).fill = mk_fill(GRAY_HDR)
    ws.cell(r, 1).font = mk_font(size=8)
    for i, run in enumerate(runs):
        ws.cell(r, 2 + i).value = run.get("project_name", "—")
        ws.cell(r, 2 + i).font = mk_font(size=8)
        ws.cell(r, 2 + i).alignment = mk_align("center")
    r += 1

    # Sub-row: run date
    ws.cell(r, 1).value = "Run date"
    ws.cell(r, 1).fill = mk_fill(GRAY_HDR)
    ws.cell(r, 1).font = mk_font(size=8)
    for i, run in enumerate(runs):
        completed = run.get("completed_at") or ""
        ws.cell(r, 2 + i).value = completed[:10] if completed else "—"
        ws.cell(r, 2 + i).font = mk_font(size=8)
        ws.cell(r, 2 + i).alignment = mk_align("center")
    r += 1

    # Sub-row: compliance
    ws.cell(r, 1).value = "Permit compliance"
    ws.cell(r, 1).fill = mk_fill(GRAY_HDR)
    ws.cell(r, 1).font = mk_font(size=8, bold=True)
    for i, run in enumerate(runs):
        summary = (run.get("results") or {}).get("summary") or {}
        compliant = summary.get("compliant")
        if compliant is True:
            ws.cell(r, 2 + i).value = "✓ PASS"
            ws.cell(r, 2 + i).font = mk_font(bold=True, color=EMERALD, size=9)
            ws.cell(r, 2 + i).fill = mk_fill(EME_LT)
        elif compliant is False:
            ws.cell(r, 2 + i).value = "✗ FAIL"
            ws.cell(r, 2 + i).font = mk_font(bold=True, color=RED, size=9)
            ws.cell(r, 2 + i).fill = mk_fill(RED_LT)
        else:
            ws.cell(r, 2 + i).value = "—"
        ws.cell(r, 2 + i).alignment = mk_align("center")
        ws.cell(r, 2 + i).border = mk_border()
    r += 2

    # ── Effluent quality table ────────────────────────────────────────────────
    ws.merge_cells(f"A{r}:{get_column_letter(2 + n)}{r}")
    ws.cell(r, 1).value = "EFFLUENT QUALITY"
    ws.cell(r, 1).font = mk_font(bold=True, color=WHITE, size=10)
    ws.cell(r, 1).fill = mk_fill(BRAND)
    ws.cell(r, 1).alignment = mk_align("center")
    r += 1

    ws.cell(r, 1).value = "Parameter (unit)"
    ws.cell(r, 1).font = mk_font(bold=True, size=9, color=WHITE)
    ws.cell(r, 1).fill = mk_fill(BRAND)
    ws.cell(r, 1).border = mk_border()
    for i, run in enumerate(runs):
        label = run.get("label") or run.get("flowsheet_name", f"Run {i+1}")
        ws.cell(r, 2 + i).value = label[:20]
        style_header_row(ws, r, 2 + i, 2 + i, BRAND)
    r += 1

    for qi, p in enumerate(QUALITY_PARAMS):
        alt = qi % 2 == 1
        style_data_row(ws, r, 1, 1 + n, alt)
        ws.cell(r, 1).value = f"{p['label']} ({p['unit']})"
        ws.cell(r, 1).font = mk_font(size=9, bold=True)

        eff_vals = []
        for i, run in enumerate(runs):
            summary = (run.get("results") or {}).get("summary") or {}
            eff = summary.get("effluent") or {}
            val = eff.get(p["key"])
            style_data_row(ws, r, 2 + i, 2 + i, alt)
            ws.cell(r, 2 + i).value = fmt_num(val, p["dec"])
            ws.cell(r, 2 + i).alignment = mk_align("right")
            ws.cell(r, 2 + i).border = mk_border()
            eff_vals.append(float(val) if val is not None else None)

        # Highlight best (lowest) effluent value for pollutants
        if p["key"] not in ("Q", "DO", "pH", "temp") and any(v is not None for v in eff_vals):
            valid = [(i, v) for i, v in enumerate(eff_vals) if v is not None]
            if len(valid) > 1:
                best_i = min(valid, key=lambda x: x[1])[0]
                ws.cell(r, 2 + best_i).font = mk_font(bold=True, color=EMERALD)
                ws.cell(r, 2 + best_i).fill = mk_fill(EME_LT)
        r += 1

    # ── Removal efficiencies ──────────────────────────────────────────────────
    r += 1
    ws.merge_cells(f"A{r}:{get_column_letter(2 + n)}{r}")
    ws.cell(r, 1).value = "REMOVAL EFFICIENCIES"
    ws.cell(r, 1).font = mk_font(bold=True, color=WHITE, size=10)
    ws.cell(r, 1).fill = mk_fill(EMERALD)
    ws.cell(r, 1).alignment = mk_align("center")
    r += 1

    rem_params = [p for p in QUALITY_PARAMS if p["key"] in ("BOD", "COD", "TSS", "TN", "NH4", "TP")]
    ws.cell(r, 1).value = "Parameter"
    ws.cell(r, 1).font = mk_font(bold=True, size=9, color=WHITE)
    ws.cell(r, 1).fill = mk_fill(EMERALD)
    ws.cell(r, 1).border = mk_border()
    for i, run in enumerate(runs):
        label = run.get("label") or run.get("flowsheet_name", f"Run {i+1}")
        ws.cell(r, 2 + i).value = label[:20]
        ws.cell(r, 2 + i).font = mk_font(bold=True, size=9, color=WHITE)
        ws.cell(r, 2 + i).fill = mk_fill(EMERALD)
        ws.cell(r, 2 + i).alignment = mk_align("center")
        ws.cell(r, 2 + i).border = mk_border()
    r += 1

    for qi, p in enumerate(rem_params):
        alt = qi % 2 == 1
        style_data_row(ws, r, 1, 1 + n, alt)
        ws.cell(r, 1).value = f"{p['label']} removal (%)"
        ws.cell(r, 1).font = mk_font(size=9, bold=True)

        rem_vals = []
        for i, run in enumerate(runs):
            summary = (run.get("results") or {}).get("summary") or {}
            inf_v = (summary.get("influent") or {}).get(p["key"])
            eff_v = (summary.get("effluent") or {}).get(p["key"])
            rem = fmt_pct(inf_v, eff_v)
            style_data_row(ws, r, 2 + i, 2 + i, alt)
            ws.cell(r, 2 + i).value = rem
            ws.cell(r, 2 + i).alignment = mk_align("right")
            ws.cell(r, 2 + i).border = mk_border()
            rem_vals.append(rem if isinstance(rem, (int, float)) else None)

        # Highlight highest removal
        valid = [(i, v) for i, v in enumerate(rem_vals) if v is not None]
        if len(valid) > 1:
            best_i = max(valid, key=lambda x: x[1])[0]
            ws.cell(r, 2 + best_i).font = mk_font(bold=True, color=EMERALD)
            ws.cell(r, 2 + best_i).fill = mk_fill(EME_LT)
        r += 1

    # ── Cost comparison sheet ────────────────────────────────────────────────
    wc = wb.create_sheet("Cost Comparison")
    wc.sheet_view.showGridLines = False
    wc.merge_cells(f"A1:{get_column_letter(2 + n)}1")
    wc["A1"].value = "Annual Operating Cost Comparison"
    wc["A1"].font = mk_font(bold=True, color=WHITE, size=13)
    wc["A1"].fill = mk_fill(NAVY)
    wc["A1"].alignment = mk_align("center")

    r = 3
    wc.cell(r, 1).value = "Cost Category"
    wc.cell(r, 1).font = mk_font(bold=True, size=9, color=WHITE)
    wc.cell(r, 1).fill = mk_fill(BRAND)
    wc.cell(r, 1).border = mk_border()
    for i, run in enumerate(runs):
        label = run.get("label") or run.get("flowsheet_name", f"Run {i+1}")
        wc.cell(r, 2 + i).value = label[:20]
        wc.cell(r, 2 + i).font = mk_font(bold=True, size=9, color=WHITE)
        wc.cell(r, 2 + i).fill = mk_fill(BRAND)
        wc.cell(r, 2 + i).alignment = mk_align("center")
        wc.cell(r, 2 + i).border = mk_border()
    r += 1

    cost_rows = [
        ("Total OPEX (USD/yr)", lambda c: c.get("total_USD_yr")),
        ("Unit cost (USD/m³)",  lambda c: c.get("cost_per_m3_treated_USD")),
        ("Energy (USD/yr)",     lambda c: (c.get("energy") or {}).get("cost_USD_yr")),
        ("Chemicals (USD/yr)",  lambda c: (c.get("chemicals") or {}).get("total_USD_yr")),
        ("Sludge (USD/yr)",     lambda c: (c.get("sludge") or {}).get("cost_USD_yr")),
        ("Labour (USD/yr)",     lambda c: (c.get("labour") or {}).get("cost_USD_yr")),
        ("Maintenance (USD/yr)",lambda c: (c.get("maintenance") or {}).get("cost_USD_yr")),
        ("Energy (kWh/yr)",     lambda c: (c.get("energy") or {}).get("total_kWh_yr")),
        ("Dry sludge (t/yr)",   lambda c: (c.get("sludge") or {}).get("dry_tonnes_yr")),
    ]

    for ci_row, (lbl, getter) in enumerate(cost_rows):
        alt = ci_row % 2 == 1
        style_data_row(wc, r, 1, 1 + n, alt)
        wc.cell(r, 1).value = lbl
        wc.cell(r, 1).font = mk_font(size=9, bold=(ci_row < 2))
        vals = []
        for i, run in enumerate(runs):
            cost = (run.get("results") or {}).get("costBreakdown")
            val = getter(cost) if cost else None
            dec = 3 if "m³" in lbl else 0
            style_data_row(wc, r, 2 + i, 2 + i, alt)
            wc.cell(r, 2 + i).value = fmt_num(val, dec) if val is not None else "—"
            wc.cell(r, 2 + i).alignment = mk_align("right")
            wc.cell(r, 2 + i).border = mk_border()
            vals.append(float(val) if val is not None else None)

        # Highlight best cost (lowest unit cost / total)
        if ci_row < 2:
            valid = [(i, v) for i, v in enumerate(vals) if v is not None]
            if len(valid) > 1:
                best_i = min(valid, key=lambda x: x[1])[0]
                wc.cell(r, 2 + best_i).font = mk_font(bold=True, color=EMERALD)
                wc.cell(r, 2 + best_i).fill = mk_fill(EME_LT)
        r += 1

    set_col_widths(ws, {"A": 30, **{get_column_letter(2 + i): 18 for i in range(n)}})
    set_col_widths(wc, {"A": 26, **{get_column_letter(2 + i): 18 for i in range(n)}})

    # ── Legend sheet ─────────────────────────────────────────────────────────
    wl = wb.create_sheet("Legend")
    wl["A1"].value = "Legend"
    wl["A1"].font = mk_font(bold=True, size=12)
    notes = [
        ("Green highlight", "Best-performing scenario for this metric"),
        ("✓ PASS",          "Effluent meets permit limit for this parameter"),
        ("✗ FAIL",          "Effluent exceeds permit limit for this parameter"),
        ("—",               "Value not available or not applicable"),
        ("Removal (%)",     "((Influent − Effluent) / Influent) × 100"),
        ("Unit cost",       "Total annual OPEX divided by annual treated volume"),
    ]
    for ri, (k, v) in enumerate(notes):
        wl.cell(2 + ri, 1).value = k
        wl.cell(2 + ri, 1).font = mk_font(bold=True, size=9)
        wl.cell(2 + ri, 2).value = v
        wl.cell(2 + ri, 2).font = mk_font(size=9)
    wl.column_dimensions["A"].width = 20
    wl.column_dimensions["B"].width = 50

    footer_r = 2 + len(notes) + 2
    wl.cell(footer_r, 1).value = f"Generated by WaterSim Pro on {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC"
    wl.cell(footer_r, 1).font = mk_font(size=8, color="888888")


# ════════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ════════════════════════════════════════════════════════════════════════════════

def main():
    raw = sys.stdin.read()
    payload = json.loads(raw)
    mode = payload.get("mode", "single")

    wb = openpyxl.Workbook()

    if mode == "comparison":
        build_comparison(wb, payload["runs"])
    else:
        build_single(wb, payload["data"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    sys.stdout.buffer.write(buf.read())


if __name__ == "__main__":
    main()
